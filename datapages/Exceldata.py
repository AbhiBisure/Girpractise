import openpyxl

book = openpyxl.load_workbook("C:\\Users\\91758\\OneDrive\\Documents\\ABEXCEL.xlsx")
# List all sheet names
sheet_names = book.sheetnames
print(sheet_names)
# Access the active sheet (default sheet)
#sheet =book.active
# Access a sheet by name
sheet = book['Sheet2']

Dic = {}
for i in range (1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value== 1:
        for j in range(2,sheet.max_column+1):
            Dic[sheet.cell(row=1,column=j).value]= sheet.cell(row=i,column=j).value


print(Dic)

